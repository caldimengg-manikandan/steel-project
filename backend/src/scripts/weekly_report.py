import os
import sys
import time
import schedule
import pymongo
from datetime import datetime
import datetime as dt_module
from dotenv import load_dotenv
import smtplib
import json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

# Load environment variables
load_dotenv()

# --- Configuration ---
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017/steel_dms")
EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASS")
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))
RECIPIENT_EMAIL = "thejathangavel05@gmail.com"

# Path to logo for Excel branding
LOGO_PATH = r"c:\Users\Arshad Ibrahim\steel-project\frontend\src\assets\logo\caldim_engineering_logo.jpg"

def get_project_stats():
    """Fetch project statistics using aggregation, scoped to the user's admin ID."""
    from bson import ObjectId
    client = pymongo.MongoClient(MONGO_URI)
    db = client.get_database()
    
    # Target Admin ID for the live projects (admin1)
    ADMIN_ID = ObjectId("699d4924d9dfdefb578dce14")
    
    # 1. Fetch live projects for this admin
    projects = list(db.projects.find({"createdByAdminId": ADMIN_ID, "status": "active"}))
    if not projects:
        client.close()
        return []
        
    project_ids = [p['_id'] for p in projects]
    stats = []

    # 2. Aggregate Drawing Stats (Matching adminProjectsController.js)
    pipeline_drawings = [
        {"$match": {"projectId": {"$in": project_ids}}},
        {"$group": {
            "_id": "$projectId",
            "totalCount": {"$sum": 1},
            "approvalCount": {
                "$sum": {
                    "$cond": [
                        {
                            "$or": [
                                {"$regexMatch": {"input": {"$ifNull": ["$extractedFields.revision", ""]}, "regex": r"^(rev\s*)?[a-z]", "options": "i"}},
                                {"$regexMatch": {"input": {"$ifNull": ["$extractedFields.remarks", ""]}, "regex": r"approved|approval", "options": "i"}},
                                {"$regexMatch": {"input": {"$ifNull": ["$extractedFields.description", ""]}, "regex": r"approved|approval", "options": "i"}}
                            ]
                        }, 1, 0
                    ]
                }
            },
            "fabricationCount": {
                "$sum": {
                    "$cond": [
                        {"$regexMatch": {"input": {"$ifNull": ["$extractedFields.revision", ""]}, "regex": r"^(rev\s*)?[0-9]", "options": "i"}},
                        1, 0
                    ]
                }
            }
        }}
    ]
    drawing_stats = {res['_id']: res for res in db.drawing_extractions.aggregate(pipeline_drawings)}

    # 3. Aggregate RFI Stats
    pipeline_rfis = [
        {"$match": {"projectId": {"$in": project_ids}}},
        {"$unwind": "$rfis"},
        {"$group": {
            "_id": "$projectId",
            "openCount": {"$sum": {"$cond": [{"$eq": ["$rfis.status", "OPEN"]}, 1, 0]}},
            "closedCount": {"$sum": {"$cond": [{"$eq": ["$rfis.status", "CLOSED"]}, 1, 0]}}
        }}
    ]
    rfi_stats = {res['_id']: res for res in db.rfiextractions.aggregate(pipeline_rfis)}

    # 4. Assemble Results
    for proj in projects:
        p_id = proj['_id']
        d_stat = drawing_stats.get(p_id, {})
        r_stat = rfi_stats.get(p_id, {})
        
        total = d_stat.get("totalCount", 0)
        approved = d_stat.get("approvalCount", 0)
        fabricated = d_stat.get("fabricationCount", 0)
        approx = proj.get("approximateDrawingsCount") or 0
        
        # Calculate percentages based on approximate count as per UI
        approval_pct = round((approved / approx * 100), 1) if approx > 0 else 0
        fab_pct = round((fabricated / approx * 100), 1) if approx > 0 else 0

        stats.append({
            "name": proj.get("name", "Unknown"),
            "client": proj.get("clientName", "Unknown"),
            "total_drawings": total,
            "approved": approved,
            "fabricated": fabricated,
            "open_rfis": r_stat.get("openCount", 0),
            "closed_rfis": r_stat.get("closedCount", 0),
            "approval_pct": approval_pct,
            "fab_pct": fab_pct
        })
        
    client.close()
    return stats

def generate_report_excel(stats):
    """Generate a styled Excel report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Status Report"
    
    # 1. Branding Header
    if os.path.exists(LOGO_PATH):
        img = Image(LOGO_PATH)
        img.width = 120
        img.height = 45
        ws.add_image(img, 'A1')
    
    ws['C2'] = "PROJECT STATUS WEEKLY REPORT"
    ws['C2'].font = Font(size=18, bold=True, color="1F4E78")
    ws['C2'].alignment = Alignment(horizontal="center")
    
    ws['C3'] = f"Generated Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['C3'].font = Font(italic=True)
    ws['C3'].alignment = Alignment(horizontal="center")
    
    # 2. Table Headers
    headers = [
        "Project Name", "Client", "Total Drawings", "Approved", "Fabricated", 
        "Approval %", "Fabrication %", "Open RFIs", "Closed RFIs"
    ]
    
    header_row = 6
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # 3. Data Rows
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for i, p in enumerate(stats, 1):
        row = header_row + i
        ws.cell(row=row, column=1, value=p['name']).border = thin_border
        ws.cell(row=row, column=2, value=p['client']).border = thin_border
        ws.cell(row=row, column=3, value=p['total_drawings']).border = thin_border
        ws.cell(row=row, column=4, value=p['approved']).border = thin_border
        ws.cell(row=row, column=5, value=p['fabricated']).border = thin_border
        
        # Pcts with conditional coloring
        app_cell = ws.cell(row=row, column=6, value=f"{p['approval_pct']}%")
        app_cell.border = thin_border
        if p['approval_pct'] >= 90: app_cell.font = Font(color="00B050", bold=True)
        
        fab_cell = ws.cell(row=row, column=7, value=f"{p['fab_pct']}%")
        fab_cell.border = thin_border
        if p['fab_pct'] >= 90: fab_cell.font = Font(color="00B050", bold=True)
        
        # RFI coloring
        open_cell = ws.cell(row=row, column=8, value=p['open_rfis'])
        open_cell.border = thin_border
        if p['open_rfis'] > 0: open_cell.font = Font(color="FF0000", bold=True)
        
        ws.cell(row=row, column=9, value=p['closed_rfis']).border = thin_border

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    filename = f"Project_Status_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename

def send_email(filename):
    """Send email with attachment."""
    if not EMAIL_USER or not EMAIL_PASS:
        print("Error: EMAIL_USER or EMAIL_PASS not set in environment.")
        return False
        
    msg = MIMEMultipart()
    msg['From'] = EMAIL_USER
    msg['To'] = RECIPIENT_EMAIL
    msg['Subject'] = f"Weekly Project Status Report - {datetime.now().strftime('%d %b %Y')}"
    
    body = f"""
    Dear Detailing Team,
    
    Please find attached the automated Weekly Project Status Report for the steel detailing projects.
    
    Summary of Active Projects: {datetime.now().strftime('%Y-%m-%d')}
    
    This is an automated report generated by the Steel DMS system.
    """
    msg.attach(MIMEText(body, 'plain'))
    
    with open(filename, "rb") as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {filename}")
        msg.attach(part)
        
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.send_message(msg)
        server.quit()
        print(f"Successfully sent report to {RECIPIENT_EMAIL}")
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

def job():
    print(f"Starting scheduled report job at {datetime.now()}")
    try:
        stats = get_project_stats()
        if not stats:
            print("No active projects found. Skipping report.")
            return
        filename = generate_report_excel(stats)
        success = send_email(filename)
        if success:
            os.remove(filename)  # Clean up after sending
    except Exception as e:
        print(f"Error in report job: {e}")

def run_scheduler():
    # Schedule the job for TUESDAY at 12:00 PM
    schedule.every().tuesday.at("12:00").do(job)
    
    print(f"Scheduler started. Waiting for Tuesday 12:00 PM... (Current time: {datetime.now().strftime('%A %H:%M')})")
    
    while True:
        schedule.run_pending()
        time.sleep(60)

if __name__ == "__main__":
    if "--now" in sys.argv:
        job()
    else:
        run_scheduler()

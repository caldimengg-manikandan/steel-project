"""
============================================================
Client Service  (Python equivalent of controllers/adminClientsController.js)
============================================================
All client CRUD + bulk Excel upload.

Routes:
  GET    /api/admin/clients
  POST   /api/admin/clients
  POST   /api/admin/clients/bulk     (Excel/CSV file upload)
  PATCH  /api/admin/clients/{id}
  DELETE /api/admin/clients/{id}
============================================================
"""

import io
import re
import logging
from datetime import datetime, timezone
from typing import Optional

from bson import ObjectId
from fastapi import HTTPException, UploadFile
from pydantic import BaseModel

from database import get_db

logger = logging.getLogger("client_service")


# ── Pydantic Schemas ──────────────────────────────────────

class ContactSchema(BaseModel):
    name: str
    email: str
    phone: Optional[str] = ""
    designation: Optional[str] = ""


class CreateClientBody(BaseModel):
    name: str
    contacts: Optional[list[ContactSchema]] = []
    status: Optional[str] = "active"


class UpdateClientBody(BaseModel):
    name: Optional[str] = None
    contacts: Optional[list] = None
    status: Optional[str] = None


# ── Helpers ──────────────────────────────────────────────

def _safe_oid(val: str) -> ObjectId:
    try:
        return ObjectId(val)
    except Exception:
        raise HTTPException(status_code=400, detail=f"Invalid ID format: {val}")


def _serialize(obj):
    """Recursively convert ObjectIds and nested structures to ensure JSON serialization."""
    if isinstance(obj, list):
        return [_serialize(i) for i in obj]
    elif isinstance(obj, dict):
        res = {}
        for k, v in obj.items():
            if k == '_id':
                res['id'] = str(v)
                res['_id'] = str(v)
            else:
                res[k] = _serialize(v)
        return res
    if isinstance(obj, ObjectId):
        return str(obj)
    return obj


# ── CRUD Operations ──────────────────────────────────────

async def list_clients(principal: dict) -> dict:
    """
    GET /api/admin/clients
    Returns all clients sorted by name.
    Mirrors listClients() in adminClientsController.js.
    """
    db = get_db()
    cursor = db["clients"].find({}).sort("name", 1)
    clients = await cursor.to_list(None)
    serialized = [_serialize(c) for c in clients]
    return {"count": len(serialized), "clients": serialized}


async def create_client(principal: dict, body: CreateClientBody) -> dict:
    """
    POST /api/admin/clients
    Mirrors createClient() in adminClientsController.js.
    """
    db = get_db()
    admin_id = principal["adminId"]

    if not body.name:
        raise HTTPException(status_code=400, detail="Client name is required.")

    now = datetime.now(timezone.utc)
    doc = {
        "name": body.name.strip(),
        "contacts": [c.model_dump() for c in (body.contacts or [])],
        "status": body.status or "active",
        "createdByAdminId": ObjectId(admin_id),
        "createdAt": now,
        "updatedAt": now,
    }

    try:
        result = await db["clients"].insert_one(doc)
    except Exception as e:
        if "duplicate key" in str(e).lower() or "11000" in str(e):
            raise HTTPException(status_code=400, detail="A client with this name already exists.")
        raise HTTPException(status_code=500, detail=str(e))

    client = await db["clients"].find_one({"_id": result.inserted_id})
    return {"client": _serialize(client)}


async def update_client(principal: dict, client_id: str, body: UpdateClientBody) -> dict:
    """
    PATCH /api/admin/clients/{client_id}
    Mirrors updateClient() in adminClientsController.js.
    """
    db = get_db()
    admin_id = principal["adminId"]

    client = await db["clients"].find_one({
        "_id": _safe_oid(client_id),
        "createdByAdminId": ObjectId(admin_id),
    })
    if not client:
        raise HTTPException(status_code=404, detail="Client not found.")

    updates: dict = {"updatedAt": datetime.now(timezone.utc)}
    if body.name is not None:
        updates["name"] = body.name.strip()
    if body.contacts is not None:
        updates["contacts"] = body.contacts
    if body.status is not None:
        updates["status"] = body.status

    await db["clients"].update_one({"_id": client["_id"]}, {"$set": updates})
    updated = await db["clients"].find_one({"_id": client["_id"]})
    return {"client": _serialize(updated)}


async def delete_client(principal: dict, client_id: str) -> dict:
    """
    DELETE /api/admin/clients/{client_id}
    Prevents deletion if active projects are linked.
    Mirrors deleteClient() in adminClientsController.js.
    """
    db = get_db()
    admin_id = principal["adminId"]

    client = await db["clients"].find_one({
        "_id": _safe_oid(client_id),
        "createdByAdminId": ObjectId(admin_id),
    })
    if not client:
        raise HTTPException(status_code=404, detail="Client not found.")

    # Block deletion if projects are linked
    linked_count = await db["projects"].count_documents({"clientId": client["_id"]})
    if linked_count > 0:
        raise HTTPException(status_code=400, detail="Cannot delete client with active projects.")

    await db["clients"].delete_one({"_id": client["_id"]})
    return {"message": f"Client \"{client['name']}\" deleted successfully."}


async def bulk_create_clients(principal: dict, file: UploadFile) -> dict:
    """
    POST /api/admin/clients/bulk
    Parse an Excel (.xlsx) or CSV file and create clients in batch.
    Mirrors bulkCreateClients() in adminClientsController.js.

    Required columns: 'Client Name', 'Client Email'
    Optional columns: 'Contact Name', 'Phone'
    """
    import openpyxl
    import csv

    db = get_db()
    admin_id = principal["adminId"]

    if not file:
        raise HTTPException(status_code=400, detail="No file uploaded.")

    content = await file.read()
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    rows = []

    try:
        if ext in ("xlsx", "xls"):
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active

            # Map headers
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            col_map = {}
            for i, h in enumerate(headers):
                if "client name" in h or h in ("name", "company", "company name"):
                    col_map["clientName"] = i
                elif "email" in h or "mail id" in h:
                    col_map["email"] = i
                elif "contact name" in h or h == "contact":
                    col_map["contactName"] = i
                elif "phone" in h or "mobile" in h:
                    col_map["phone"] = i

            if "clientName" not in col_map:
                raise HTTPException(status_code=400, detail='Could not find mandatory "Client Name" column.')
            if "email" not in col_map:
                raise HTTPException(status_code=400, detail='Could not find mandatory "Client Email" column.')

            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(row)

        elif ext == "csv":
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.reader(text.splitlines())
            header_row = next(reader, [])
            headers = [h.strip().lower() for h in header_row]
            col_map = {}
            for i, h in enumerate(headers):
                if "client name" in h or h in ("name", "company"):
                    col_map["clientName"] = i
                elif "email" in h:
                    col_map["email"] = i
                elif "contact name" in h or h == "contact":
                    col_map["contactName"] = i
                elif "phone" in h or "mobile" in h:
                    col_map["phone"] = i

            if "clientName" not in col_map:
                raise HTTPException(status_code=400, detail='Could not find mandatory "Client Name" column.')
            if "email" not in col_map:
                raise HTTPException(status_code=400, detail='Could not find mandatory "Client Email" column.')

            for row in reader:
                rows.append(tuple(row))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Upload .xlsx or .csv")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse file: {e}")

    created_count = 0
    skipped_count = 0
    error_list = []

    for i, row in enumerate(rows, start=2):
        try:
            client_name_raw = row[col_map["clientName"]] if len(row) > col_map["clientName"] else None
            email_raw = row[col_map["email"]] if len(row) > col_map["email"] else None
            contact_name_raw = row[col_map.get("contactName", -1)] if col_map.get("contactName") is not None and len(row) > col_map["contactName"] else None
            phone_raw = row[col_map.get("phone", -1)] if col_map.get("phone") is not None and len(row) > col_map["phone"] else None

            if not client_name_raw or not email_raw:
                error_list.append(f"Row {i}: Missing Client Name or Email.")
                skipped_count += 1
                continue

            name = str(client_name_raw).strip()
            email = str(email_raw).strip()
            contact_name = str(contact_name_raw).strip() if contact_name_raw else name
            phone = str(phone_raw).strip() if phone_raw else ""

            # Duplicate check (case-insensitive)
            existing = await db["clients"].find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}})
            if existing:
                error_list.append(f"Row {i}: Client \"{name}\" already exists.")
                skipped_count += 1
                continue

            now = datetime.now(timezone.utc)
            await db["clients"].insert_one({
                "name": name,
                "contacts": [{"name": contact_name, "email": email, "phone": phone, "designation": ""}],
                "status": "active",
                "createdByAdminId": ObjectId(admin_id),
                "createdAt": now,
                "updatedAt": now,
            })
            created_count += 1

        except Exception as e:
            error_list.append(f"Row {i}: {str(e)}")
            skipped_count += 1

    return {
        "message": f"Bulk upload completed. Created: {created_count}, Skipped: {skipped_count}.",
        "createdCount": created_count,
        "skippedCount": skipped_count,
        "errors": error_list,
    }

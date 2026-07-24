import json
import re
from datetime import datetime

def is_col3(t):
    # validate Date format: MM-DD-YYYY, etc.
    return bool(re.search(r'\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}', t))

def is_col0(t):
    # numbers [1 to N number]
    return t.isdigit()

def is_col1(t):
    # any number or any alphabet max of two digit or character
    return (t.isdigit() or t.isalpha()) and len(t) <= 2

def is_col2(t):
    # a sentences or a word - atleast 3 character
    return len(t) >= 3 and not is_col3(t)

def parse_date(date_str):
    match = re.search(r'\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}', date_str)
    if not match: return datetime.min
    ds = match.group()
    for fmt in ('%m-%d-%Y', '%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(ds, fmt)
        except ValueError:
            pass
    return datetime.min

def assign_token(row_dict, t):
    if is_col3(t):
        if 3 in row_dict: return False
        row_dict[3] = t
        return True
    
    if is_col2(t):
        if 2 in row_dict: return False
        row_dict[2] = t
        return True

    if is_col0(t) and 0 not in row_dict:
        row_dict[0] = t
        return True
    
    if is_col1(t) and 1 not in row_dict:
        row_dict[1] = t
        return True
        
    return False

def process_revision_table(rows):
    if not rows: return rows
    
    remove_keywords_rev = ["#", "# of", "Copies", "# of Copies", "Revision", "Issue", "Revision/Issue", "Destination", "Date", "REVISION/"]
    
    all_tokens = []
    for row in rows:
        for cell in row:
            all_tokens.extend([t.strip() for t in cell.split('\n') if t.strip()])
            
    filtered_tokens = []
    for t in all_tokens:
        tu = t.upper().strip()
        is_header = False
        for kw in remove_keywords_rev:
            if kw.upper() == tu:
                is_header = True
                break
        if not is_header:
            filtered_tokens.append(t)
            
    final_rows = []
    row_dict = {}
    for t in filtered_tokens:
        assigned = assign_token(row_dict, t)
        if not assigned:
            if row_dict:
                final_rows.append([row_dict.get(0, ""), row_dict.get(1, ""), row_dict.get(2, ""), row_dict.get(3, "")])
                row_dict = {}
                assign_token(row_dict, t)
                
    if row_dict:
        final_rows.append([row_dict.get(0, ""), row_dict.get(1, ""), row_dict.get(2, ""), row_dict.get(3, "")])
        
    # Sort the Rows according to the latest date
    final_rows.sort(key=lambda x: parse_date(x[3]), reverse=True)
    
    return final_rows

def process_simple_field(rows, keywords):
    if not rows: return rows
    new_rows = []
    for row in rows:
        keep = True
        for cell in row:
            cu = cell.upper().strip()
            for kw in keywords:
                if kw.upper() in cu:
                    keep = False
                    break
            if not keep: break
        if keep:
            new_rows.append(row)
    return new_rows

def generate_excel(data, output_excel="extraction_results.xlsx"):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        print("openpyxl is not installed. Please install it using 'pip install openpyxl'")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drawings"

    # Define styles
    green_font = Font(color="008000", bold=True)
    red_font = Font(color="FF0000", bold=True)
    black_bold = Font(bold=True)
    title_font = Font(size=20, bold=True)

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells('B2:F2')
    ws['B2'] = "CALDIM ENGINEERING PRIVATE LIMITED"
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

    # Project Info
    ws.merge_cells('A4:C4')
    ws['A4'] = "PROJECT NAME : AAAA"
    ws['A4'].font = green_font
    
    ws.merge_cells('E4:F4')
    ws['E4'] = "TRANSMITTAL NO: #333"
    ws['E4'].font = green_font
    ws['E4'].alignment = Alignment(horizontal='right')

    project_no = "N/A"
    for filename, detections in data.items():
        for det in detections:
            if det.get("label") == "PROJECT_NO" and det.get("rows"):
                project_no = " ".join(det["rows"][0])
                break
        if project_no != "N/A":
            break

    ws.merge_cells('A5:C5')
    ws['A5'] = f"PROJECT NO : {project_no}"
    ws['A5'].font = green_font

    from datetime import datetime
    today_str = datetime.now().strftime("%m-%d-%Y")
    ws.merge_cells('E5:F5')
    ws['E5'] = f"Date: {today_str}"
    ws['E5'].font = green_font
    ws['E5'].alignment = Alignment(horizontal='right')

    ws.merge_cells('A6:C6')
    ws['A6'] = "FABRICATOR : SSSS"
    ws['A6'].font = green_font

    # Headers
    headers = ["Sl. No.", "DrawingNo.", "Drawing Description", "REV#", "DATE", "Remarks"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num)
        cell.value = header
        cell.font = black_bold
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # DRAWINGS separator
    ws.merge_cells('A9:F9')
    cell = ws['A9']
    cell.value = "DRAWINGS"
    cell.font = red_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col_num in range(1, 7):
        ws.cell(row=9, column=col_num).border = border

    # Data Rows
    row_num = 10
    sl_no = 1
    for filename, detections in data.items():
        drawing_no = ""
        drawing_desc = ""
        rev = ""
        date_str = ""
        remarks = ""
        
        for det in detections:
            label = det.get("label", "")
            rows = det.get("rows", [])
            if not rows: continue
            
            if label == "DRAWING_NO":
                drawing_no = " ".join(rows[0])
            elif label == "DRAWING_DESCRIPTION":
                drawing_desc = " \n ".join([" ".join(r) for r in rows])
            elif label == "REVISION_TABLE":
                latest_row = rows[0]
                if len(latest_row) >= 4:
                    col0 = latest_row[0]
                    col1 = latest_row[1]
                    rev = col1 if col1 else col0
                    remarks = latest_row[2]
                    date_str = latest_row[3]

        if not drawing_no and not drawing_desc:
            continue
            
        row_data = [sl_no, drawing_no, drawing_desc, rev, date_str, remarks]
        for col_index, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_index)
            cell.value = val
            cell.border = border
            if col_index in [1, 4]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif col_index == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                
        sl_no += 1
        row_num += 1

    # Adjust column widths
    column_widths = {'A': 8, 'B': 20, 'C': 40, 'D': 10, 'E': 15, 'F': 30}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
        
    wb.save(output_excel)
    print(f"Excel file generated successfully with custom template: {output_excel}")

def main():
    input_file = "detection_results.json"
    output_file = "detection_results.json"
    
    with open(input_file, "r") as f:
        data = json.load(f)
        
    for filename, detections in data.items():
        for det in detections:
            label = det.get("label", "")
            rows = det.get("rows", [])
            
            if label == "DRAWING_NO":
                det["rows"] = process_simple_field(rows, ["Drawing", "#", "Drawing #"])
            elif label == "PROJECT_NO":
                det["rows"] = process_simple_field(rows, ["Contract", "#", "Contract #"])
                det["rows"] = process_simple_field(det["rows"], ["Drawing", "Title", "Drawing Title:"])
            elif label == "DRAWING_DESCRIPTION":
                cleaned_rows = process_simple_field(rows, ["Drawing", "Title", "Drawing Title:", "Sheet", "Sheet No.", "Sheet No", "Sheet No:", "SHEET", "SHEET NO."])
                if cleaned_rows:
                    concatenated_text = " \n ".join([" ".join(r) for r in cleaned_rows]).strip()
                    det["rows"] = [[concatenated_text]]
                else:
                    det["rows"] = []
            elif label == "REVISION_TABLE":
                det["rows"] = process_revision_table(rows)
                
    with open(output_file, "w") as f:
        json.dump(data, f, indent=4)
        
    print(f"Extraction formatting script completed. Processed results saved to {output_file}")
    
    generate_excel(data, "extraction_results.xlsx")

if __name__ == "__main__":
    main()

"""
============================================================
Excel Generator  (Python equivalent of services/excelService.js)
============================================================
Uses openpyxl (pure-Python, no COM dependency) to build the same
styled Excel workbooks that the Node ExcelJS version produced.

Exported functions:
  generate_project_excel(rows, project_details, type)
      → (bytes, filename)  — generate a fresh workbook in memory

  append_rows_to_project_excel(project_id, rows)
      → str — append rows to the per-project cached workbook & return path

Logo path mirrors the Node version (frontend/src/assets/excel_im/excel_img.png)
============================================================
"""

import io
import os
import re
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

logger = logging.getLogger("excel_generator")

# ── Paths ─────────────────────────────────────────────────
_BASE = Path(__file__).parent
_EXCEL_DIR = _BASE / "uploads" / "excel"
_LOGO_PATH = _BASE.parent / "frontend" / "src" / "assets" / "excel_im" / "excel_img.png"
_EXCEL_DIR.mkdir(parents=True, exist_ok=True)

THIN = Side(style="thin")
MEDIUM = Side(style="medium")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MEDIUM_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def _fill(argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb)


def _font(bold=False, size=10, color="FF000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic)


def _align(h="center", v="middle", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _style(cell, bold=False, size=10, color="FF000000", bg=None,
           h="center", v="middle", wrap=True, border=True, italic=False):
    cell.font = _font(bold=bold, size=size, color=color, italic=italic)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = _align(h=h, v=v, wrap=wrap)
    if border:
        cell.border = THIN_BORDER


def _merge_style(ws, min_row, min_col, max_row, max_col, **kwargs):
    ws.merge_cells(
        start_row=min_row, start_column=min_col,
        end_row=max_row, end_column=max_col
    )
    _style(ws.cell(min_row, min_col), **kwargs)


# ── Shared revision helpers (mirrors transmittalService.js) ──

def normalize_revision(rev: str) -> str:
    if not rev:
        return ""
    return re.sub(r"^rev[\s\-_]*", "", str(rev).strip(), flags=re.IGNORECASE).upper()


def _revision_rank(rev: str) -> int:
    norm = normalize_revision(rev)
    if not norm:
        return -1
    try:
        n = int(norm)
        return 10000 + n   # fabrication tier
    except ValueError:
        return ord(norm[0])  # approval tier: A=65, B=66 …


def pick_latest_revision(history: list) -> dict:
    """Return the entry with the most advanced revision mark."""
    if not history:
        return {}
    best = history[0]
    for entry in history[1:]:
        if _revision_rank(entry.get("mark", "")) > _revision_rank(best.get("mark", "")):
            best = entry
    return best


# ── Logo helper ────────────────────────────────────────────

def _try_add_logo(ws, tl_col, tl_row, br_col, br_row):
    """Embed the Caldim logo if it exists on disk."""
    if _LOGO_PATH.exists():
        try:
            img = XLImage(str(_LOGO_PATH))
            # openpyxl anchors by pixel: approx 60 px per column, 18 px per row
            img.anchor = f"{get_column_letter(tl_col + 1)}{tl_row + 1}"
            img.width = (br_col - tl_col) * 64
            img.height = (br_row - tl_row) * 18
            ws.add_image(img)
        except Exception as e:
            logger.warning(f"[Excel] Logo embed failed: {e}")


# ══════════════════════════════════════════════════════════
# Main public function
# ══════════════════════════════════════════════════════════

def generate_project_excel(
    rows: list,
    project_details: dict,
    sheet_type: Optional[str] = None,
) -> tuple[bytes, str]:
    """
    Build a fresh multi-sheet Excel workbook in memory.

    Parameters
    ----------
    rows           : list of DrawingExtraction dicts (MongoDB .to_list() output)
    project_details: {"projectName": ..., "clientName": ..., "transmittalNo": ...}
    sheet_type     : "transmittal" | "log" | None (= both sheets)

    Returns
    -------
    (buffer_bytes, filename)
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet
    wb.creator = "Caldim Steel Detailing DMS"

    project_name = str(project_details.get("projectName") or "Project")
    client_name  = str(project_details.get("clientName") or "UNKNOWN")
    transmittal_no = int(project_details.get("transmittalNo") or 1)

    today = datetime.now()
    formatted_date = today.strftime("%m/%d/%Y")

    # ── Helper functions shared between sheets ────────────
    def _get_latest_rev(f: dict) -> str:
        hist = f.get("revisionHistory") or []
        if hist:
            return pick_latest_revision(hist).get("mark") or f.get("revision") or ""
        return f.get("revision") or ""

    def _get_latest_date(f: dict) -> str:
        hist = f.get("revisionHistory") or []
        if hist:
            return pick_latest_revision(hist).get("date") or f.get("date") or ""
        return f.get("date") or ""

    def _get_latest_remarks(f: dict) -> str:
        hist = f.get("revisionHistory") or []
        if hist:
            return pick_latest_revision(hist).get("remarks") or f.get("remarks") or ""
        return f.get("remarks") or ""

    # ══════════════════════════════════════════════════════
    # SHEET 1 — Transmittal
    # ══════════════════════════════════════════════════════
    if not sheet_type or sheet_type == "transmittal":
        tr = wb.create_sheet("Transmittal")

        # Logo rows 1-6
        for r in range(1, 7):
            tr.row_dimensions[r].height = 18
        tr.row_dimensions[7].height = 6
        _try_add_logo(tr, 0, 0, 5, 6)

        T = 8  # data starts at row 8

        green_font = _font(bold=True, size=12, color="FF00B050")

        # Row T: Project name | Transmittal no
        tr.row_dimensions[T].height = 22
        c = tr.cell(T, 1, f"PROJECT NAME : {project_name.upper()}")
        c.font = green_font
        tr.merge_cells(start_row=T, start_column=1, end_row=T, end_column=3)
        c2 = tr.cell(T, 4, f"TRANSMITTAL NO: TR-{str(transmittal_no).zfill(3)}")
        c2.font = green_font
        c2.alignment = _align(h="right")
        tr.merge_cells(start_row=T, start_column=4, end_row=T, end_column=6)

        # Row T+1: Fabricator | Date
        tr.row_dimensions[T + 1].height = 22
        c = tr.cell(T + 1, 1, f"FABRICATOR   : {client_name.upper()}")
        c.font = green_font
        tr.merge_cells(start_row=T + 1, start_column=1, end_row=T + 1, end_column=3)
        c2 = tr.cell(T + 1, 4, f"DATE: {formatted_date}")
        c2.font = green_font
        c2.alignment = _align(h="right")
        tr.merge_cells(start_row=T + 1, start_column=4, end_row=T + 1, end_column=6)

        tr.row_dimensions[T + 2].height = 8  # spacer

        # Column widths
        tr.column_dimensions["A"].width = 12
        tr.column_dimensions["B"].width = 22
        tr.column_dimensions["C"].width = 50
        tr.column_dimensions["D"].width = 18
        tr.column_dimensions["E"].width = 18
        tr.column_dimensions["F"].width = 40

        # Detect if fabrication data present for dynamic header
        has_fab = any(
            (lambda rev: rev and not re.match(r'^[a-zA-Z]+$', rev))(
                (r.get("extractedFields") or {}).get("revision", "")
            )
            for r in rows
        )
        header_label = "Sent for Fabrication" if has_fab else "Sent for Approval"

        # Header rows T+3 and T+4
        tr.row_dimensions[T + 3].height = 24
        tr.row_dimensions[T + 4].height = 22
        hfill = _fill("FFD9E1F2")
        hfont = _font(bold=True, size=10, color="FF1F3864")

        def _th(row, col, val):
            c = tr.cell(row, col, val)
            c.font = hfont
            c.fill = hfill
            c.alignment = _align()
            c.border = THIN_BORDER

        for col in range(1, 7):
            _th(T + 3, col, "")
            _th(T + 4, col, "")

        tr.cell(T + 3, 1).value = "Sl. No."
        tr.cell(T + 3, 2).value = "Sheet No."
        tr.cell(T + 3, 3).value = "Drawing Title"
        tr.cell(T + 3, 4).value = header_label
        tr.cell(T + 3, 6).value = "Revision History"
        tr.cell(T + 4, 4).value = "REV#"
        tr.cell(T + 4, 5).value = "DATE"

        # Vertical merges for non-split columns
        for col in [1, 2, 3, 6]:
            tr.merge_cells(start_row=T + 3, start_column=col,
                           end_row=T + 4, end_column=col)
        # Horizontal merge for header label
        tr.merge_cells(start_row=T + 3, start_column=4,
                       end_row=T + 3, end_column=5)

        tr.freeze_panes = f"A{T + 5}"

        # Data rows — grouped by folderName
        folder_groups: dict[str, list] = {}
        for r in rows:
            fn = r.get("folderName") or "DETAIL SHEETS"
            folder_groups.setdefault(fn, []).append(r)

        sl_no = 1
        for folder in sorted(folder_groups.keys()):
            group = sorted(
                folder_groups[folder],
                key=lambda r: (r.get("extractedFields") or {}).get("drawingNumber") or r.get("originalFileName") or ""
            )

            # Folder header row (yellow)
            fr = tr.max_row + 1
            tr.row_dimensions[fr].height = 22
            c = tr.cell(fr, 1, folder.upper())
            c.font = _font(bold=True, size=11)
            c.fill = _fill("FFFFFF00")
            c.alignment = _align()
            c.border = THIN_BORDER
            tr.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=6)
            for col in range(2, 7):
                cell = tr.cell(fr, col)
                cell.fill = _fill("FFFFFF00")
                cell.border = THIN_BORDER

            # Drawing rows
            for dr in group:
                f = dr.get("extractedFields") or {}
                rev_hist = f.get("revisionHistory") or []
                hist_str = " | ".join(
                    f"Rev {h.get('mark', '')} ({h.get('date', '')})"
                    for h in rev_hist if h.get("mark")
                ) or _get_latest_rev(f)

                rn = tr.max_row + 1
                tr.row_dimensions[rn].height = 22
                data = [
                    sl_no,
                    f.get("drawingNumber") or "",
                    f.get("drawingTitle") or f.get("drawingDescription") or dr.get("originalFileName") or "",
                    _get_latest_rev(f),
                    _get_latest_date(f),
                    hist_str,
                ]
                for ci, val in enumerate(data, start=1):
                    cell = tr.cell(rn, ci, val)
                    cell.border = THIN_BORDER
                    cell.alignment = _align(
                        h="left" if ci in (3, 6) else "center"
                    )
                sl_no += 1

    # ══════════════════════════════════════════════════════
    # SHEET 2 — Drawing Log
    # ══════════════════════════════════════════════════════
    if not sheet_type or sheet_type == "log":
        lg = wb.create_sheet("Drawing Log")

        for r in range(1, 7):
            lg.row_dimensions[r].height = 18
        lg.row_dimensions[7].height = 6
        _try_add_logo(lg, 0, 0, 2, 5)

        L = 8

        # ── Build revision column lists ────────────────────
        all_revs: set[str] = set()
        for row in rows:
            f = row.get("extractedFields") or {}
            hist = f.get("revisionHistory") or [{"mark": f.get("revision"), "remarks": f.get("remarks")}]
            for h in hist:
                mark = str(h.get("mark") or "").upper().strip()
                rem = str(h.get("remarks") or "").lower()
                if mark == "0" and "fabrication" not in rem:
                    continue
                if mark:
                    all_revs.add(mark)

        alpha_revs = sorted([r for r in all_revs if re.match(r'^[A-Z]$', r)])
        num_revs = sorted([r for r in all_revs if re.match(r'^\d+$', r)], key=lambda x: int(x))
        if "A" not in alpha_revs:
            alpha_revs.insert(0, "A")

        total_revs = len(alpha_revs) + len(num_revs)
        total_cols = max(3 + total_revs + 1, 4)

        # ── Title row L ────────────────────────────────────
        lg.row_dimensions[L].height = 28
        _merge_style(lg, L, 1, L, total_cols,
                     bold=True, size=14, bg="FFFFF2CC", h="center", v="middle", wrap=False)
        lg.cell(L, 1).value = "OUTGOING DRAWING LOG SHEET"

        # ── Project name + client L+1 ──────────────────────
        lg.row_dimensions[L + 1].height = 24
        mid = (total_cols // 2)
        _merge_style(lg, L + 1, 1, L + 1, mid, bold=True, size=11, h="left", wrap=False)
        lg.cell(L + 1, 1).value = f"Project Name : {project_name}"
        _merge_style(lg, L + 1, mid + 1, L + 1, total_cols, bold=True, size=11, h="left", wrap=False)
        lg.cell(L + 1, mid + 1).value = f"Client : {client_name}"

        # ── Group header L+2 and sub-header L+3 ───────────
        lg.row_dimensions[L + 2].height = 24
        lg.row_dimensions[L + 3].height = 22

        approval_fill = "FFB4C6E7"
        fabric_fill   = "FFC6E0B4"
        grey_fill     = "FFD9D9D9"
        hfont = _font(bold=True, size=10, color="FF1F3864")

        def _lh(row, col, val, bg=grey_fill):
            c = lg.cell(row, col, val)
            c.font = hfont
            c.fill = _fill(bg)
            c.alignment = _align()
            c.border = THIN_BORDER

        for idx, label in enumerate(["Sl. No", "Sheet No", "Drawing Title"]):
            col = idx + 1
            _lh(L + 2, col, label)
            _lh(L + 3, col, label)
            lg.merge_cells(start_row=L + 2, start_column=col,
                           end_row=L + 3, end_column=col)

        cur_col = 4
        if alpha_revs:
            _lh(L + 2, cur_col, "Sent for Approval", bg=approval_fill)
            if len(alpha_revs) > 1:
                lg.merge_cells(start_row=L + 2, start_column=cur_col,
                               end_row=L + 2, end_column=cur_col + len(alpha_revs) - 1)
            for r in alpha_revs:
                _lh(L + 2, cur_col, "Sent for Approval", bg=approval_fill)
                _lh(L + 3, cur_col, f"Rev {r}", bg=approval_fill)
                lg.column_dimensions[get_column_letter(cur_col)].width = 14
                cur_col += 1

        if num_revs:
            _lh(L + 2, cur_col, "Sent for Fabrication", bg=fabric_fill)
            if len(num_revs) > 1:
                lg.merge_cells(start_row=L + 2, start_column=cur_col,
                               end_row=L + 2, end_column=cur_col + len(num_revs) - 1)
            for r in num_revs:
                _lh(L + 2, cur_col, "Sent for Fabrication", bg=fabric_fill)
                _lh(L + 3, cur_col, f"Rev {r}", bg=fabric_fill)
                lg.column_dimensions[get_column_letter(cur_col)].width = 14
                cur_col += 1

        remarks_col = cur_col
        _lh(L + 2, remarks_col, "Remarks")
        _lh(L + 3, remarks_col, "Remarks")
        lg.merge_cells(start_row=L + 2, start_column=remarks_col,
                       end_row=L + 3, end_column=remarks_col)

        lg.column_dimensions["A"].width = 10
        lg.column_dimensions["B"].width = 22
        lg.column_dimensions["C"].width = 45
        lg.column_dimensions[get_column_letter(remarks_col)].width = 40
        lg.freeze_panes = f"A{L + 4}"

        # ── DRAWINGS section label ─────────────────────────
        sec_row = lg.max_row + 1
        lg.row_dimensions[sec_row].height = 22
        _merge_style(lg, sec_row, 1, sec_row, total_cols,
                     bold=True, size=11, bg="FFFFFF00")
        lg.cell(sec_row, 1).value = "DRAWINGS"

        # ── Consolidate revisions per drawing number ───────
        group_dwg: dict = {}
        for row in rows:
            f = row.get("extractedFields") or {}
            d_num = f.get("drawingNumber") or "UNKNOWN"
            if d_num not in group_dwg:
                group_dwg[d_num] = {
                    "drawingNumber": d_num,
                    "drawingTitle": f.get("drawingTitle") or f.get("drawingDescription") or row.get("originalFileName") or "",
                    "revisions": [],
                    "remarks_list": [],
                }
            hist = f.get("revisionHistory") or [{"mark": f.get("revision"), "date": f.get("date"), "remarks": f.get("remarks")}]
            for h in hist:
                if h.get("mark") not in (None, ""):
                    group_dwg[d_num]["revisions"].append({
                        "mark": str(h.get("mark", "")).upper().strip(),
                        "date": h.get("date") or "",
                        "remarks": h.get("remarks") or "",
                    })
            if f.get("remarks"):
                group_dwg[d_num]["remarks_list"].append(str(f["remarks"]).upper().strip())

        dwgs_sorted = sorted(
            group_dwg.values(),
            key=lambda d: d["drawingNumber"],
        )

        log_sl = 1
        for d in dwgs_sorted:
            rev_map = {rev["mark"]: rev["date"] for rev in d["revisions"] if rev["mark"]}
            combined_remarks = " / ".join(sorted(set(filter(None, d["remarks_list"]))))
            has_num_rev = any(rev_map.get(r) for r in num_revs)
            alpha_start_col = 4
            alpha_end_col   = 3 + len(alpha_revs)

            row_data = [log_sl, d["drawingNumber"], d["drawingTitle"]]
            for r in alpha_revs:
                row_data.append(rev_map.get(r) or "")
            for r in num_revs:
                row_data.append(rev_map.get(r) or "")
            row_data.append(combined_remarks)

            dr = lg.max_row + 1
            lg.row_dimensions[dr].height = 22

            for ci, val in enumerate(row_data, start=1):
                cell = lg.cell(dr, ci, val)
                cell.border = THIN_BORDER
                cell.alignment = _align(h="left" if ci in (3, remarks_col) else "center")

                is_alpha_col = alpha_start_col <= ci <= alpha_end_col
                is_rev_a_col = (ci == alpha_start_col)

                # Highlight blank Rev A cell if drawing already has fabrication (skipped approval)
                if is_rev_a_col and not val and has_num_rev:
                    cell.fill = _fill("FFECECEC")

            log_sl += 1

    # ── Write to buffer ────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    safe_name = re.sub(r"[^\w\-]", "_", project_name)
    if sheet_type == "log":
        filename = f"{safe_name}_Drawing_Log.xlsx"
    else:
        filename = f"{safe_name}_Transmittal.xlsx"

    return raw, filename


# ══════════════════════════════════════════════════════════
# Batch append helper (used by extraction_pipeline.py)
# ══════════════════════════════════════════════════════════

def append_rows_to_project_excel(project_id: str, rows: list) -> str:
    """
    Append rows to the per-project cached Excel file.
    Creates the file with headers if it doesn't exist.
    Returns the absolute path to the file.
    Mirrors appendRowsToProjectExcel() in excelService.js.
    """
    if not rows:
        return None

    file_path = _EXCEL_DIR / f"{project_id}_drawings.xlsx"

    COLUMNS = [
        ("Sl. No.", 8),
        ("Sheet No.", 22),
        ("Drawing Title", 45),
        ("Revision Mark", 14),
        ("Date", 16),
        ("Remarks", 40),
        ("Original Filename", 30),
    ]

    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb["Drawing Log"] if "Drawing Log" in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Drawing Log"
        wb.creator = "Steel Detailing DMS"

        # Headers
        for ci, (hdr, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(1, ci, hdr)
            cell.font = _font(bold=True)
            cell.fill = _fill("FFFFFFFF")
            cell.alignment = _align()
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(ci)].width = width

        ws.row_dimensions[1].height = 30

        # Sub-header "SHOP DRAWING"
        ws.append(["SHOP DRAWING"] + [""] * (len(COLUMNS) - 1))
        ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
        sub_cell = ws.cell(2, 1)
        sub_cell.font = _font(bold=True, size=12)
        sub_cell.fill = _fill("FFFFFF00")
        sub_cell.alignment = _align()
        sub_cell.border = THIN_BORDER
        ws.row_dimensions[2].height = 25
        ws.freeze_panes = "A3"

    start_sl = max(ws.max_row - 1, 0)

    for idx, row in enumerate(rows):
        sl_no = max(start_sl + idx, 1)
        ws.append([
            sl_no,
            row.get("drawingNumber") or "",
            row.get("drawingTitle") or row.get("drawingDescription") or "",
            row.get("revision") or "",
            row.get("date") or "",
            row.get("remarks") or "",
            row.get("fileName") or "",
        ])
        nr = ws.max_row
        ws.row_dimensions[nr].height = 22
        for ci in range(1, len(COLUMNS) + 1):
            cell = ws.cell(nr, ci)
            cell.border = THIN_BORDER
            cell.alignment = _align()

    wb.save(file_path)
    logger.info(f"[Excel] Batch saved {len(rows)} rows → {file_path}")
    return str(file_path)
